import openpyxl
import csv
import re

def load_replacements(csv_file):
    replacements = {}
    with open(csv_file, 'r', newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            find_value = row['current']
            replacements[find_value] = True  # Store exact case from the CSV (for lookup purposes)
    return replacements

def exact_case_match(text, find_value):
    """
    Check if the find_value matches exactly in the text, followed by non-alphanumeric or end of string.
    """
    pattern = r'\b' + re.escape(find_value) + r'(?=[^\w]|$)'  # \b ensures it's a whole word, (?=[^\w]|$) ensures no alphanumeric follows
    return re.search(pattern, text) is not None

def search_unreplaced_values(excel_file, csv_file, sheet_name, columns):
    replacements = load_replacements(csv_file)

    wb = openpyxl.load_workbook(excel_file)

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found.")
        return

    sheet = wb[sheet_name]
    found_mismatches = False

    for col in columns:
        for row in sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value and isinstance(cell.value, str):  # Ensure the cell contains text
                    matched = any(exact_case_match(cell.value, find_value) for find_value in replacements)
                    if not matched:
                        print(f"Unreplaced value '{cell.value}' found at {cell.coordinate}")
                        found_mismatches = True

    if not found_mismatches:
        print("All values matched with replacements.")