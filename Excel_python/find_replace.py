import openpyxl
import csv
import re

# Function to load the CSV file as a dictionary
def load_replacements(csv_file):
    replacements = {}
    with open(csv_file, 'r', newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            find_value = row['current']
            replace_value = row['replace']
            replacements[find_value] = replace_value  # Store exact case from the CSV
    return replacements

# Function to perform the exact case-sensitive replacement
def exact_case_replace(text, find_value, replace_value):
    """
    Replace the find_value in text with the replace_value only if they match exactly.
    """
    # Create a case-sensitive pattern to match the exact word followed by non-alphanumeric or end of string
    pattern = r'\b' + re.escape(find_value) + r'(?=[^\w]|$)'  # \b ensures it's a whole word, (?=[^\w]|$) ensures no alphanumeric follows
    return re.sub(pattern, replace_value, text)

# Function to replace text in a specific Excel sheet and column
def replace_in_excel(sheet, columns, replacements):
    for col in columns:
        for row in sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value and isinstance(cell.value, str):  # Ensure the cell contains text
                    original_text = cell.value
                    for find_value, replace_value in replacements.items():
                        # Perform the exact case-sensitive replacement
                        cell.value = exact_case_replace(cell.value, find_value, replace_value)

# Function to process the entire workbook
def process_workbook(excel_file, replacements, sheet_names, columns):
    # Open the workbook
    wb = openpyxl.load_workbook(excel_file)

    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            replace_in_excel(sheet, columns, replacements)
        else:
            print(f"Sheet '{sheet_name}' not found.")

    # Save the modified workbook
    wb.save("modified_" + excel_file)

# Example usage
# csv_file = 'replacements.csv'  # CSV file with current and replace columns
# excel_file = 'workbook.xlsx'   # Excel file to process
# sheet_names = ['Sheet1', 'Sheet2']  # List of sheet names to process
# columns = [1, 2]  # List of columns (e.g., column 1, 2 for columns A, B)

# # Load replacements from CSV
# replacements = load_replacements(csv_file)

# # Process the workbook and apply replacements
# process_workbook(excel_file, replacements, sheet_names, columns)